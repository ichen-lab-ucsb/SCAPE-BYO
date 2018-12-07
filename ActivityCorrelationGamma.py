# This is a python script to calculate fitness correlation values (gamma)
# by Celia Blanco
# contact: cblanco@chem.ucsb.edu

from openpyxl import load_workbook
import Levenshtein 
import sys

########## File name, sheet name and correlation distance ##########

file_name = str(sys.argv[1])
sheet_name = str(sys.argv[2])
dis = int(sys.argv[3])

wb = load_workbook(file_name) # Load in the workbook
sheet = wb[sheet_name] # Load in the sheet
print 'd = ', dis

########## Create list with sequences ('seq') and activity values ('f') ##########

list = []
count = 0
for i in range (1, sheet.max_row): 
	list.append({'seq':  str(sheet.cell(row=i+1, column=1).value), 'f': sheet.cell(row=i+1, column=4).value})
	count += 1

print 'Number of Sequences =', count
print ""

########### Create list of sequence pairs one mutation apart - store sequences, activity values and position and type of mutation ##########
########### First part of numerator in equation - also denominator ########### 

list_mut = []

num = 0
den = 0
elements = 0

for seq1 in list:
	for seq2 in list:
		if Levenshtein.distance(seq1['seq'], seq2['seq']) == 1:
			ops = Levenshtein.editops(seq1['seq'], seq2['seq'])
			for i in range(0, len(ops)) :
				pos_op = ops[i][0]
				if pos_op == 'replace':
					pos_ori = ops[i][1]
					pos_des = ops[i][2]
					let_ori = seq1['seq'][pos_ori]
					let_des = seq2['seq'][pos_des]
					pos = pos_ori
					list_mut.append({'seq1': seq1['seq'], 'seq2': seq2['seq'], 'f1': seq1['f'], 'f2': seq2['f'], 'from': let_ori, 'to': let_des, 'pos': pos, 'd': Levenshtein.distance(seq1['seq'],seq2['seq'])})
				if pos_op == 'delete':
					pos_ori = ops[i][1]
					let_ori = seq1['seq'][pos_ori]
					let_des = '-'
					pos = pos_ori
					list_mut.append({'seq1': seq1['seq'], 'seq2': seq2['seq'], 'f1': seq1['f'], 'f2': seq2['f'], 'from': let_ori, 'to': let_des, 'pos': pos, 'd': Levenshtein.distance(seq1['seq'],seq2['seq'])})
				if pos_op == 'insert':
					pos_des = ops[i][2]
					let_ori = '-'
					let_des = seq2['seq'][pos_des]
					pos = pos_des
					list_mut.append({'seq1': seq1['seq'], 'seq2': seq2['seq'], 'f1': seq1['f'], 'f2': seq2['f'], 'from': let_ori, 'to': let_des, 'pos': pos, 'd': Levenshtein.distance(seq1['seq'],seq2['seq'])})

########### Create list of sequence pairs one mutation apart - store sequences, activity values and position and type of mutation ##########
########### Second part part of numerator in equation ###########

abd = 'ACGT-'


for letter1 in abd:
	for letter2 in abd:
		for item1 in list_mut:
			if item1['from'] == letter1 and item1['to'] == letter2:
				if item1['d'] == 1:
					for item2 in list_mut:
						if item2['d'] == 1:
							if item2['from'] == letter1 and item2['to'] == letter2 and item1['pos'] == item2['pos']:
								if Levenshtein.distance(item1['seq1'],item2['seq1']) == dis:
									num += (item1['f1']-item1['f2'])*(item2['f1'] - item2['f2'])
									den += (item1['f1']-item1['f2'])*(item1['f1']-item1['f2'])
									elements += 1

					
print 'elements:', elements
print 'gamma:', float(num)/float(den)
		
